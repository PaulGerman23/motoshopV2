# apps/ventas/views_exportacion.py
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from .models import Venta, Caja


@login_required
def exportar_ventas_excel(request):
    """Exportar ventas a Excel"""
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['N° Venta', 'Fecha', 'Cliente', 'Total', 'Método Pago', 'Estado', 'Caja']
    ws.append(headers)
    
    # Aplicar estilo a headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Datos
    ventas = Venta.objects.select_related('cliente', 'caja').all()
    
    for venta in ventas:
        ws.append([
            venta.numero_venta,
            venta.fecha.strftime('%d/%m/%Y %H:%M'),
            venta.cliente.nombre,
            float(venta.total),
            venta.get_metodo_pago_display(),
            venta.get_estado_display(),
            f"Caja #{venta.caja.numero}" if venta.caja else "Sin caja"
        ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def exportar_ventas_pdf(request):
    """Exportar listado de ventas a PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph("Listado de Ventas", title_style))
    elements.append(Spacer(1, 0.5*inch))
    
    # Datos
    ventas = Venta.objects.select_related('cliente', 'caja').all()
    
    data = [['N° Venta', 'Fecha', 'Cliente', 'Total', 'Estado', 'Caja']]
    
    for venta in ventas:
        data.append([
            venta.numero_venta,
            venta.fecha.strftime('%d/%m/%Y'),
            venta.cliente.nombre,
            f'${venta.total:.2f}',
            venta.get_estado_display(),
            f"#{venta.caja.numero}" if venta.caja else "N/A"
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[1*inch, 1.2*inch, 2*inch, 1*inch, 1*inch, 1*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response


@login_required
def exportar_venta_pdf(request, venta_id):
    """Exportar una venta individual a PDF"""
    venta = get_object_or_404(Venta, id=venta_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=venta_{venta.numero_venta}.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph(f"Venta #{venta.numero_venta}", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Información general
    info_data = [
        ['Fecha:', venta.fecha.strftime('%d/%m/%Y %H:%M')],
        ['Cliente:', venta.cliente.nombre],
        ['Método de Pago:', venta.get_metodo_pago_display()],
        ['Estado:', venta.get_estado_display()],
    ]
    
    if venta.caja:
        info_data.append(['Caja:', f"#{venta.caja.numero}"])
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E9EB')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Productos
    elements.append(Paragraph("Productos", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))
    
    productos_data = [['Producto', 'Cantidad', 'Precio Unit.', 'Subtotal']]
    
    for detalle in venta.detalles.all():
        productos_data.append([
            detalle.producto.nombre,
            str(detalle.cantidad),
            f'${detalle.precio_unitario:.2f}',
            f'${detalle.subtotal:.2f}'
        ])
    
    productos_table = Table(productos_data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
    productos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(productos_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Totales
    totales_data = [
        ['Subtotal:', f'${venta.subtotal:.2f}'],
        ['IVA (21%):', f'${venta.iva:.2f}'],
        ['TOTAL:', f'${venta.total:.2f}']
    ]
    
    totales_table = Table(totales_data, colWidths=[4.5*inch, 2*inch])
    totales_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 14),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#4472C4')),
    ]))
    
    elements.append(totales_table)
    
    doc.build(elements)
    return response


@login_required
def exportar_caja_pdf(request, caja_id):
    """Exportar resumen de caja a PDF"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=caja_{caja.numero}.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#10B981') if caja.estado == 'abierta' else colors.HexColor('#6B7280'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph(f"Caja #{caja.numero}", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Información de la caja
    info_data = [
        ['Usuario:', caja.usuario.get_full_name() or caja.usuario.username],
        ['Estado:', caja.get_estado_display()],
        ['Fecha Apertura:', caja.fecha_apertura.strftime('%d/%m/%Y %H:%M')],
    ]
    
    if caja.estado == 'cerrada':
        info_data.append(['Fecha Cierre:', caja.fecha_cierre.strftime('%d/%m/%Y %H:%M')])
    
    info_data.extend([
        ['Monto Inicial:', f'${caja.monto_inicial:.2f}'],
    ])
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E9EB')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Resumen de ventas
    elements.append(Paragraph("Resumen de Ventas", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))
    
    resumen_data = [
        ['Concepto', 'Monto'],
        ['Total Ventas', f'${caja.total_ventas:.2f}'],
        ['Total Efectivo', f'${caja.total_efectivo:.2f}'],
        ['Total Transferencia', f'${caja.total_transferencia:.2f}'],
        ['Cantidad de Ventas', str(caja.cantidad_ventas)],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[4*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 1), (1, -1), 12),
    ]))
    
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Si la caja está cerrada, mostrar información de cierre
    if caja.estado == 'cerrada':
        elements.append(Paragraph("Información de Cierre", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        cierre_data = [
            ['Concepto', 'Monto'],
            ['Monto Esperado', f'${caja.monto_esperado:.2f}'],
            ['Monto Real', f'${caja.monto_real:.2f}'],
            ['Diferencia', f'${caja.diferencia:.2f}'],
        ]
        
        cierre_table = Table(cierre_data, colWidths=[4*inch, 2*inch])
        cierre_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 1), (1, -1), 12),
        ]))
        
        elements.append(cierre_table)
    
    # Lista de ventas
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("Detalle de Ventas", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))
    
    ventas_data = [['N° Venta', 'Fecha', 'Cliente', 'Método', 'Total']]
    
    for venta in caja.ventas.all():
        ventas_data.append([
            venta.numero_venta,
            venta.fecha.strftime('%d/%m/%Y %H:%M'),
            venta.cliente.nombre[:20],  # Limitar longitud
            venta.get_metodo_pago_display(),
            f'${venta.total:.2f}'
        ])
    
    ventas_table = Table(ventas_data, colWidths=[1*inch, 1.5*inch, 2*inch, 1.2*inch, 1*inch])
    ventas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(ventas_table)
    
    doc.build(elements)
    return response